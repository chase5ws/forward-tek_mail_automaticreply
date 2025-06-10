import openpyxl
from datetime import datetime  # 用於生成檔案名稱

# 讀取使用者提供的 Excel 檔案
input_file = input("請輸入要處理的 Excel 檔案名稱（包含副檔名，例如 data.xlsx）：")

try:
    # 開啟使用者提供的 Excel 檔案
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active  # 使用第一個工作表

    # 讀取資料
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳過標題列
        data.append(list(row))  # 將每一列轉成列表，加入 data 中

    # 定義 Excel 標題
    headers = ["EMAIL", "名稱", "離職日期", "狀態", "輸出內容"]

    # 建立新的 Excel 檔案
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "處理結果"

    # 寫入標題
    new_ws.append(headers)

    # 處理資料並寫入新的 Excel 檔案
    for row in data:
        email, name, leave_date, status = row
        output_content = ""  # 初始化輸出內容

        # 移除 email 中 @ 及其後的部分
        email_username = email.split('@')[0] if email else ""

        # 根據狀態生成對應的輸出內容
        if status == "轉移":
            output_content = (
                f"{email} 此為舊信箱，已變更為「@forward-tek.com.tw」 / {email} has been changed to 「@forward-tek.com.tw」\n\n"
                f"您好，\n感謝您的來信。\n本信箱所屬同仁 {email} 此為舊信箱，已變更為「@forward-tek.com.tw」\n\n謝謝您的配合與理解。\n\n"
                f"Hello,\nThank you for your message.\nThis mailbox belongs to {email}. This is the old mailbox and has been changed to 「@forward-tek.com.tw」.\n\nThank you for your understanding.\n"
            )
        elif status == "停用":
            output_content = (
                f"{name} 帳號已停用 / {email_username} account has been closed\n\n"
                f"您好，\n感謝您的來信。\n本信箱所屬同仁 {name} 帳號已停用，無法再處理郵件事宜。\n若您有相關業務需求，請改聯絡相關窗口。\n\n謝謝您的配合與理解。\n\n"
                f"Hello,\nThank you for your message.\nPlease be informed that {email_username} has close account, and this mailbox is no longer monitored.\nFor further assistance, please contact the relevant window instead.\n\nThank you for your understanding.\n"
            )
        elif status == "離職":
            output_content = (
                f"{name} 已離職 / {email_username} has left the company\n\n"
                f"您好，\n感謝您的來信。\n本信箱所屬同仁 {name} 已於 {leave_date} 離職，無法再處理郵件事宜。\n若您有相關業務需求，請改聯絡相關窗口。\n\n謝謝您的配合與理解。\n\n"
                f"Hello,\nThank you for your message.\nPlease be informed that {email_username} has left the company as of {leave_date}, and this mailbox is no longer monitored.\nFor further assistance, please contact the relevant window instead.\n\nThank you for your understanding.\n"
            )

        # 將處理後的資料寫入新的 Excel 檔案
        new_ws.append([email, name, leave_date, status, output_content])

    # 生成檔案名稱（加入當前日期和時間）
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")  # 格式化日期時間為 YYYYMMDD_HHMMSS
    output_file = f"處理完成資料_{current_time}.xlsx"  # 檔案名稱

    # 儲存新的 Excel 檔案
    new_wb.save(output_file)
    print(f"資料已成功處理並寫入 {output_file}")

except FileNotFoundError:
    print(f"無法找到檔案：{input_file}，請確認檔案名稱是否正確。")
except Exception as e:
    print(f"處理檔案時發生錯誤：{e}")
